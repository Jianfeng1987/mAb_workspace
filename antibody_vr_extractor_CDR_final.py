#!/usr/bin/env python3
"""
抗体可变区序列提取工具 + 进化树分析[命名规则：80-2-B1-96A01-H/L-001]
===================================
完整流程：FASTA/FASTQ → IgBLAST → CDR提取 → VH/VL配对 → 进化树

使用方法：
  # 交互式运行（默认）
  python antibody_vr_extractor_CDR.py /workspace/80-2.fasta

  # 命令行模式（非交互式）
  python antibody_vr_extractor_CDR.py --input /workspace/80-2.fasta \\
      --output_dir /workspace/antibody_work --cdr_def kabat \\
      --species mouse --pair --tree --skip_setup
"""

import subprocess, os, sys, io, re, tempfile, argparse
from pathlib import Path
from collections import Counter

# 确保 conda 库路径在 LD_LIBRARY_PATH 中（IgBLAST 需要 libgomp.so.1）
_conda_lib = os.path.expanduser('~/miniconda3/lib')
if _conda_lib not in os.environ.get('LD_LIBRARY_PATH', ''):
    os.environ['LD_LIBRARY_PATH'] = _conda_lib + ':' + os.environ.get('LD_LIBRARY_PATH', '')
# 确保 conda bin 路径在 PATH 中（hmmscan, muscle）
_conda_bin = os.path.expanduser('~/miniconda3/bin')
if _conda_bin not in os.environ.get('PATH', ''):
    os.environ['PATH'] = _conda_bin + ':' + os.environ.get('PATH', '')
import numpy as np
from Bio import SeqIO
from Bio.Seq import Seq
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════
#  全局配置
# ═══════════════════════════════════════════════════════════════════
WORK_DIR    = '/home/liuyunping/antibody/antibody_work'
IGBLAST_DIR = f'{WORK_DIR}/ncbi-igblast-1.22.0'
DB_DIR      = f'{WORK_DIR}/igblast_db'
IGBLASTN    = f'{IGBLAST_DIR}/bin/igblastn'

LOCAL_IGBLAST_TAR = '/home/liuyunping/antibody/ncbi-igblast-1.22.0-x64-linux.tar.gz'
LOCAL_MOUSE_DB_TAR = '/home/liuyunping/antibody/mouse_gl_VDJ.tar'

# ── 这些配置在 set_config() 中更新 ──
INPUT_FILE = None
ORGANISM   = 'mouse'
CDR_DEF   = 'kabat'       # kabat | imgt | chothia | contact | martin
# IgBLAST 原生支持的 CDR 定义（可直接通过 -domain_system 获得）
CDR_IGBLAST_NATIVE = {'kabat', 'imgt'}
USE_IMGT_NAMING = False   # 是否使用 IMGT 标准命名（IGHV1-2*01）
OUTPUT_XLSX = None
AUX_FILE   = None
VR_DNA_COLS = ['fwr1', 'cdr1', 'fwr2', 'cdr2', 'fwr3', 'cdr3', 'fwr4']
VR_AA_COLS  = ['fwr1_aa', 'cdr1_aa', 'fwr2_aa', 'cdr2_aa', 'fwr3_aa', 'cdr3_aa', 'fwr4_aa']
MIN_VH_NT = 270
MIN_VL_NT = 240

# ── CDR 相关的列名（用于配对输出）──
CDR_COLS = ['cdr1_dna', 'cdr1_aa', 'cdr1_length',
            'cdr2_dna', 'cdr2_aa', 'cdr2_length',
            'cdr3_dna', 'cdr3_aa', 'cdr3_length']

# ── COS 全局变量（在 set_config 中初始化）──
COS_CLIENT = None
COS_BUCKET = None
COS_PREFIX = None
COS_ENABLED = False


# ═══════════════════════════════════════════════════════════════════
#  SECTION 0 — 环境安装
# ═══════════════════════════════════════════════════════════════════
def step0_install():
    """Step 0: 安装 Python 依赖、HMMER、MUSCLE、IgBLAST 和鼠源数据库"""
    print("=" * 60)
    print("  Step 0 — 安装环境")
    print("=" * 60)

    # 1. Python 依赖
    print("\n[1/5] 检查 Python 依赖...")
    for mod in ['Bio', 'pandas', 'openpyxl', 'numpy']:
        try:
            __import__(mod)
        except ImportError:
            print(f"  ⚠️ 缺少 {mod}，请运行: pip install {mod}")
    print("  ✓ 依赖已就绪")

    # 1.5 HMMER（abnumber/ANARCI 所需）
    print("\n[2/6] 检查 HMMER...")
    if subprocess.run(['which', 'hmmscan'], capture_output=True).returncode != 0:
        subprocess.run(['conda', 'install', '-y', '-c', 'bioconda', 'hmmer'], check=False)
        if subprocess.run(['which', 'hmmscan'], capture_output=True).returncode != 0:
            print("  ⚠️ HMMER 未安装，请运行: conda install -c bioconda hmmer")
        else:
            print("  ✓ HMMER 安装完成")
    else:
        print("  ✓ HMMER 已存在")

    # 1.6 MUSCLE（进化树对齐所需）
    print("\n[3/6] 检查 MUSCLE...")
    if subprocess.run(['which', 'muscle'], capture_output=True).returncode != 0:
        subprocess.run(['conda', 'install', '-y', '-c', 'bioconda', 'muscle'], check=False)
        if subprocess.run(['which', 'muscle'], capture_output=True).returncode != 0:
            print("  ⚠️ MUSCLE 未安装，请运行: conda install -c bioconda muscle")
        else:
            print("  ✓ MUSCLE 安装完成")
    else:
        print("  ✓ MUSCLE 已存在")

    # 2. 创建工作目录
    os.makedirs(WORK_DIR, exist_ok=True)
    os.makedirs(DB_DIR, exist_ok=True)

    # 3. IgBLAST
    print("\n[4/6] 安装 IgBLAST 1.22.0...")
    if not os.path.exists(f'{IGBLAST_DIR}/bin/igblastn'):
        if os.path.exists(LOCAL_IGBLAST_TAR):
            subprocess.run(f'tar -xzf {LOCAL_IGBLAST_TAR} -C {WORK_DIR}/', shell=True, check=True)
            subprocess.run(f'chmod +x {IGBLAST_DIR}/bin/igblastn', shell=True)
            print("  ✓ IgBLAST 安装完成")
        else:
            print(f"  ❌ 找不到本地文件: {LOCAL_IGBLAST_TAR}")
            print("     请确认文件已上传到 /home/liuyunping/antibody/")
            return False
    else:
        print("  ✓ IgBLAST 已存在")

    # 4. 鼠源数据库
    print("\n[5/6] 安装鼠源 IMGT 数据库...")
    if not os.path.exists(f'{DB_DIR}/mouse_gl_V.nhr'):
        if os.path.exists(LOCAL_MOUSE_DB_TAR):
            subprocess.run(f'tar -xf {LOCAL_MOUSE_DB_TAR} -C {DB_DIR}/', shell=True, check=True)
            print("  ✓ 数据库安装完成")
        else:
            print(f"  ❌ 找不到本地文件: {LOCAL_MOUSE_DB_TAR}")
            return False
    else:
        print("  ✓ 数据库已存在")

    # 验证
    print("\n[6/6] 验证...")
    result = subprocess.run([IGBLASTN, '-version'], capture_output=True, text=True)
    if result.returncode != 0:
        print(f"  ⚠️ IgBLAST 无法运行: {result.stderr.strip()[:200]}")
        print(f"  提示: 确保 LD_LIBRARY_PATH 包含 libgomp.so.1 所在目录")
        print(f"  当前 LD_LIBRARY_PATH={os.environ.get('LD_LIBRARY_PATH', '')}")
    else:
        version_line = result.stdout.strip().split('\n')[0]
        print(f"  ✓ IgBLAST 版本: {version_line}")
    db_files = sum(1 for f in Path(DB_DIR).glob('mouse_gl_*') if f.is_file())
    print(f"  ✓ 数据库文件数: {db_files}")
    print("\n  ✅ 环境准备完成！\n")
    return True


# ═══════════════════════════════════════════════════════════════════
#  IMGT 数据库设置（标准命名，如 IGHV1-2*01）
# ═══════════════════════════════════════════════════════════════════
IMGT_DB_DIR = f'{WORK_DIR}/imgt_db'

def setup_imgt_database():
    """从 IMGT 官网下载 Mus musculus 种系序列，构建标准命名数据库"""
    print("\n" + "=" * 60)
    print("  [IMGT] 下载并构建 IMGT 标准命名数据库")
    print("  [IMGT] 基因名格式: IGHV1-2*01, IGHV3-23*01, IGHJ4*01 ...")
    print("=" * 60)

    makeblastdb = f'{IGBLAST_DIR}/bin/makeblastdb'
    edit_imgt   = f'{IGBLAST_DIR}/bin/edit_imgt_file.pl'
    imgt_raw_dir = f'{IMGT_DB_DIR}/imgt_raw'
    os.makedirs(imgt_raw_dir, exist_ok=True)

    imgt_base = 'https://www.imgt.org/download/V-QUEST/IMGT_V-QUEST_reference_directory/Mus_musculus/IG/'
    imgt_files = {
        'IGHV': 'IGHV.fasta', 'IGHD': 'IGHD.fasta', 'IGHJ': 'IGHJ.fasta',
        'IGKV': 'IGKV.fasta', 'IGKJ': 'IGKJ.fasta',
        'IGLV': 'IGLV.fasta', 'IGLJ': 'IGLJ.fasta',
    }

    # 下载
    print("\n  [1/3] 下载 IMGT Mus musculus 种系序列...")
    all_ok = True
    for gene, fname in imgt_files.items():
        dst = os.path.join(imgt_raw_dir, fname)
        if os.path.exists(dst) and os.path.getsize(dst) > 100:
            print(f"    ✓ {gene}: 已存在")
            continue
        url = imgt_base + fname
        r = subprocess.run(f'wget -q "{url}" -O "{dst}"', shell=True,
                           capture_output=True, text=True)
        if r.returncode != 0 or os.path.getsize(dst) < 100:
            subprocess.run(f'curl -s "{url}" -o "{dst}"', shell=True)
        if os.path.exists(dst) and os.path.getsize(dst) > 100:
            print(f"    ✓ {gene}: {os.path.getsize(dst)//1024}KB")
        else:
            print(f"    ✗ {gene}: 下载失败，请检查网络")
            all_ok = False
    if not all_ok:
        print("\n  ⚠️ 部分文件下载失败，IMGT 命名不可用")
        return False

    # 处理 header + 建库
    print("\n  [2/3] 处理 IMGT header，合并 V/D/J 文件...")
    merge_map = {
        'mouse_imgt_ig_v': ['IGHV.fasta', 'IGKV.fasta', 'IGLV.fasta'],
        'mouse_imgt_ig_d': ['IGHD.fasta'],
        'mouse_imgt_ig_j': ['IGHJ.fasta', 'IGKJ.fasta', 'IGLJ.fasta'],
    }
    for db_name, source_files in merge_map.items():
        merged = os.path.join(imgt_raw_dir, f'{db_name}_merged.fasta')
        with open(merged, 'w') as out:
            for sf in source_files:
                src = os.path.join(imgt_raw_dir, sf)
                if os.path.exists(src) and os.path.getsize(src) > 100:
                    with open(src) as f: out.write(f.read())

        processed = os.path.join(IMGT_DB_DIR, f'{db_name}.fasta')
        if os.path.exists(edit_imgt):
            subprocess.run(f'perl "{edit_imgt}" "{merged}" > "{processed}"',
                           shell=True, capture_output=True)
        else:
            # Python 备用：提取 IMGT header 的第一个字段（基因名）
            with open(merged) as fi, open(processed, 'w') as fo:
                for line in fi:
                    if line.startswith('>'):
                        gene = line[1:].split('|')[0].strip()
                        gene = re.sub(r'\s+', '_', gene)
                        fo.write(f'>{gene}\n')
                    else:
                        fo.write(line.replace('.', '').upper())

        # 建 BLAST 数据库
        r = subprocess.run(
            f'"{makeblastdb}" -parse_seqids -dbtype nucl -in "{processed}"',
            shell=True, capture_output=True, text=True)
        if r.returncode == 0:
            print(f'    ✓ {db_name} 建库完成')
        else:
            print(f'    ✗ {db_name} 建库失败: {r.stderr[:100]}')
            return False

    print("\n  [3/3] 验证...")
    # 检查标准 BLAST DB 文件（可能带 .fasta. 前缀）
    v_path = f'{IMGT_DB_DIR}/mouse_imgt_ig_v'
    v_files = os.listdir(IMGT_DB_DIR)
    v_db_ok = any('mouse_imgt_ig_v' in f and f.endswith('.nhr') for f in v_files)
    if not v_db_ok:
        print(f"  ✗ IMGT V 数据库不完整")
        return False
    print("  ✅ IMGT 标准命名数据库就绪!")
    print(f"     数据库目录: {IMGT_DB_DIR}")
    return True


# ═══════════════════════════════════════════════════════════════════
#  SECTION 1 — 用户配置
# ═══════════════════════════════════════════════════════════════════
def set_config():
    """Step 1: 配置输入文件路径和 COS 参数"""
    global INPUT_FILE, ORGANISM, OUTPUT_XLSX, AUX_FILE
    global COS_CLIENT, COS_BUCKET, COS_PREFIX, COS_ENABLED
    global CDR_DEF, USE_IMGT_NAMING

    print("=" * 60)
    print("  Step 1 — 参数 & COS 配置")
    print("=" * 60)
    print()

    # ── 输入文件 ──────────────────────────────────────────────────────
    if len(sys.argv) > 1:
        INPUT_FILE = sys.argv[1]
        print(f"  输入文件 (来自参数): {INPUT_FILE}")
    else:
        INPUT_FILE = input("  请输入 FASTA/FASTQ 文件路径: ").strip()
        if not INPUT_FILE:
            INPUT_FILE = '/home/liuyunping/antibody/80-2.fasta'
            print(f"  使用默认: {INPUT_FILE}")

    # 物种选择
    print("\n  可选物种: mouse, human, rabbit, rat")
    org_input = input(f"  请输入物种 [默认: {ORGANISM}]: ").strip()
    if org_input:
        ORGANISM = org_input

    # CDR 定义选择
    cdr_map = {
        'kabat': 'kabat', 'k': 'kabat',
        'imgt': 'imgt',   'i': 'imgt',
        'chothia': 'chothia', 'c': 'chothia',
        'contact': 'contact',
        'martin': 'martin', 'm': 'martin',
    }
    print("\n  可选 CDR 定义:")
    print("    IgBLAST 原生: kabat (默认), imgt")
    print("    abnumber 扩展: chothia, contact, martin")
    cdr_input = input(f"  请输入 CDR 定义 [默认: {CDR_DEF}]: ").strip().lower()
    if cdr_input:
        CDR_DEF = cdr_map.get(cdr_input, cdr_input)
        if CDR_DEF not in ('kabat', 'imgt', 'chothia', 'contact', 'martin'):
            print(f"    ⚠️  无效定义 '{cdr_input}'，使用默认 kabat")
            CDR_DEF = 'kabat'

    # 是否使用 IMGT 标准命名
    print("\n  推荐使用 IMGT 标准命名（如 IGHV1-2*01, IGHJ4*01）")
    print("  需要从 IMGT 官网下载数据库（首次需网络）")
    imgt_input = input(f"  是否使用 IMGT 标准命名? (y/N): ").strip().lower()
    USE_IMGT_NAMING = (imgt_input == 'y')
    if USE_IMGT_NAMING:
        print("  ⏳ 正在下载并构建 IMGT 数据库（约 1-2 分钟）...")
        if not setup_imgt_database():
            print("  ⚠️  IMGT 数据库构建失败，使用默认命名")
            USE_IMGT_NAMING = False

    OUTPUT_XLSX = f'{WORK_DIR}/antibody_results_{CDR_DEF}.xlsx'
    AUX_FILE = f'{IGBLAST_DIR}/optional_file/{ORGANISM}_gl.aux'

    print(f"\n  配置摘要:")
    print(f"    输入文件: {INPUT_FILE}")
    print(f"    物种:     {ORGANISM}")
    print(f"    CDR 定义: {CDR_DEF.upper()}")
    print(f"    IMGT 命名: {'✅ 是' if USE_IMGT_NAMING else '否（使用 NCBI 命名）'}")
    print(f"    输出:     {OUTPUT_XLSX}")

    if os.path.exists(INPUT_FILE):
        size = os.path.getsize(INPUT_FILE)
        print(f"    文件大小: {size:,} bytes ({size/1024:.1f} KB)")
    else:
        print(f"    ⚠️  文件不存在: {INPUT_FILE}")
        return False

    # ── COS 配置 ──────────────────────────────────────────────────────
    COS_ENABLED = False
    COS_CLIENT, COS_BUCKET, COS_PREFIX = None, None, None

    use_cos = input("\n  是否使用腾讯云 COS 存储? (y/N): ").strip().lower()
    if use_cos == 'y':
        sid    = input("  SecretId: ").strip()
        skey   = input("  SecretKey: ").strip()
        region = input("  地域 (如 ap-beijing): ").strip() or 'ap-beijing'
        bucket = input("  存储桶名称: ").strip()
        prefix = input("  文件前缀 (默认 antibody_analysis/): ").strip() or 'antibody_analysis/'

        from qcloud_cos import CosConfig, CosS3Client
        config = CosConfig(Region=region, SecretId=sid, SecretKey=skey)
        COS_CLIENT = CosS3Client(config)
        COS_BUCKET = bucket
        COS_PREFIX = prefix

        # 验证连接
        try:
            COS_CLIENT.head_bucket(Bucket=bucket)
            COS_ENABLED = True
            print(f"  ✅ COS 连接成功 | {bucket}/{prefix}")
        except Exception as e:
            print(f"  ⚠️  COS 连接失败: {e}")
            print("  将继续本地分析，无需 COS")

    return True


# ═══════════════════════════════════════════════════════════════════
#  辅助函数
# ═══════════════════════════════════════════════════════════════════
def _get(row, key, default=''):
    """安全获取 Series 行中的字符串值，去除空值标记"""
    if row is None or (hasattr(row, 'empty') and row.empty):
        return default
    v = row.get(key)
    if v is None:
        return default
    v = str(v).strip()
    return default if v.lower() in ('nan', 'none', '', 'null') else v


def extract_cdrs_abnumber(vr_aa, vr_dna, cdr_def):
    """使用 abnumber (ANARCI) 按指定 CDR 定义提取 CDR1/2/3 的 AA 与 DNA 序列"""
    try:
        from abnumber import Chain
        chain = Chain(vr_aa, scheme=cdr_def)
        positions_list = list(chain.positions.keys())

        def get_cdr(cdr_dict):
            if not cdr_dict:
                return '', '', 0
            cdr_keys = set(cdr_dict.keys())
            raw_idxs = [i for i, pos in enumerate(positions_list) if pos in cdr_keys]
            if not raw_idxs:
                return '', '', 0
            start = raw_idxs[0]
            end   = raw_idxs[-1] + 1
            aa_seq = vr_aa[start:end]
            nt_start = start * 3
            nt_end   = min(end * 3, len(vr_dna))
            dna_seq = vr_dna[nt_start:nt_end]
            return dna_seq, aa_seq, len(dna_seq)

        cdr1_dna, cdr1_aa, cdr1_len = get_cdr(chain.cdr1_dict)
        cdr2_dna, cdr2_aa, cdr2_len = get_cdr(chain.cdr2_dict)
        cdr3_dna, cdr3_aa, cdr3_len = get_cdr(chain.cdr3_dict)
        return cdr1_dna, cdr1_aa, cdr1_len, cdr2_dna, cdr2_aa, cdr2_len, cdr3_dna, cdr3_aa, cdr3_len
    except Exception as e:
        print(f"    ⚠️  abnumber 编号失败 ({e})，回退到 IgBLAST 原始 CDR")
        return None


def reverse_complement(seq):
    """返回序列的反向互补链"""
    return str(Seq(seq).reverse_complement())


def read_sequences(input_file):
    """读取 FASTA 或 FASTQ 文件"""
    suffix = Path(input_file).suffix.lower()
    fmt = 'fastq' if suffix in ('.fastq', '.fq') else 'fasta'
    return [{'id': rec.id, 'seq': str(rec.seq).upper()}
            for rec in SeqIO.parse(input_file, fmt)]


def write_fasta(records, path):
    """将序列列表写入 FASTA 文件"""
    with open(path, 'w') as f:
        for r in records:
            f.write(f">{r['id']}\n{r['seq']}\n")


def run_igblast(fasta_path):
    """运行 IgBLAST，返回 AIRR TSV 格式的 DataFrame"""
    env = os.environ.copy()
    env['IGDATA'] = IGBLAST_DIR

    # 选择数据库：IMGT 标准命名 vs NCBI 默认命名
    if USE_IMGT_NAMING:
        db_v = f'{IMGT_DB_DIR}/mouse_imgt_ig_v.fasta'
        db_d = f'{IMGT_DB_DIR}/mouse_imgt_ig_d.fasta'
        db_j = f'{IMGT_DB_DIR}/mouse_imgt_ig_j.fasta'
        aux  = f'{IGBLAST_DIR}/optional_file/mouse_gl.aux'
    else:
        db_v = f'{DB_DIR}/{ORGANISM}_gl_V'
        db_d = f'{DB_DIR}/{ORGANISM}_gl_D'
        db_j = f'{DB_DIR}/{ORGANISM}_gl_J'
        aux  = AUX_FILE

    cmd = [
        IGBLASTN,
        '-germline_db_V', db_v,
        '-germline_db_D', db_d,
        '-germline_db_J', db_j,
        '-organism', ORGANISM,
        '-query', fasta_path,
        '-auxiliary_data', aux,
        '-outfmt', '19',
        '-num_threads', '1',
    ]
    # 只有 IgBLAST 原生支持的 CDR 定义才使用 -domain_system，其他由 abnumber 后处理
    if CDR_DEF in CDR_IGBLAST_NATIVE:
        cmd.extend(['-domain_system', CDR_DEF])
    result = subprocess.run(cmd, capture_output=True, text=True, env=env)
    if result.returncode != 0:
        err = result.stderr.strip()
        print(f"  ❌ IgBLAST 进程退出码 {result.returncode}")
        print(f"  🔍 IGDATA = {env.get('IGDATA', 'NOT SET')}")
        print(f"  🔍 数据库路径 = {db_v}")
        print(f"  🔍 辅助文件 = {aux}")
        # 检查 internal_data 是否存在
        internal_org = f'{IGBLAST_DIR}/internal_data/{ORGANISM}'
        if os.path.isdir(internal_org):
            files = [f for f in os.listdir(internal_org) if 'mouse_V' in f][:5]
            print(f"  🔍 internal_data/{ORGANISM} 文件(前5): {files}")
        else:
            print(f"  🔍 internal_data/{ORGANISM} 不存在!")
        raise RuntimeError(f"IgBLAST 运行失败:\n{err}")
    if not result.stdout.strip():
        print(f"  ⚠️ IgBLAST 输出为空! stderr: {result.stderr[:500]}")
        print(f"  🔍 命令: {' '.join(cmd)}")
        return pd.DataFrame()
    return pd.read_csv(io.StringIO(result.stdout), sep='\t', dtype=str,
                       keep_default_na=False)


def concat_vr(row, cols):
    """拼接可变区各区段序列，跳过点号和 NA"""
    parts = []
    for c in cols:
        val = _get(row, c)
        if val and val not in ('.', 'nan', 'None'):
            parts.append(val)
    return ''.join(parts)


def clean_cdr(val):
    """清理 CDR 字符串：去除空值和点号"""
    v = val.replace('.', '') if val else ''
    return v if v and v.lower() not in ('nan', 'none') else ''


def score_hit(row):
    """对 IgBLAST 命中质量打分"""
    if row is None:
        return 0
    v_call = _get(row, 'v_call')
    if not v_call:
        return 0
    score = 1
    if _get(row, 'productive').upper() in ('T', 'TRUE'):   score += 4
    if _get(row, 'complete_vdj').upper() in ('T', 'TRUE'): score += 2
    if _get(row, 'stop_codon').upper()   not in ('T', 'TRUE'): score += 1
    return score


def qc_check(row, vr_dna, vr_aa):
    """质控检查"""
    v_call = _get(row, 'v_call')
    if not v_call:
        return 'FAIL_NO_HIT', 'No V gene match'
    if _get(row, 'stop_codon').upper() in ('T', 'TRUE'):
        pos = vr_aa.find('*') + 1 if '*' in vr_aa else '?'
        return 'FAIL_STOP_CODON', f'Stop codon at position {pos}'
    locus = _get(row, 'locus').upper()
    min_len = MIN_VH_NT if locus == 'IGH' else MIN_VL_NT
    if len(vr_dna) < min_len:
        return 'FAIL_SHORT', f'VR length {len(vr_dna)} nt < minimum {min_len} nt'
    if not vr_dna:
        return 'FAIL_NO_VR', 'Variable region sequence is empty'
    if _get(row, 'complete_vdj').upper() not in ('T', 'TRUE'):
        return 'PASS_INCOMPLETE_J', 'J region may be truncated (Sanger read length)'
    return 'PASS', ''


# ═══════════════════════════════════════════════════════════════════
#  SECTION 2 — 主分析流程
# ═══════════════════════════════════════════════════════════════════
def analyze():
    """Step 2-7: 读取序列 → IgBLAST → 提取可变区 → 输出 Excel"""
    print("\n" + "=" * 60)
    print("  Step 2 — 读取序列")
    print("=" * 60)
    records = read_sequences(INPUT_FILE)
    n_total = len(records)
    n_heavy = sum(1 for r in records if '-H-' in r['id'])
    n_light = sum(1 for r in records if '-L-' in r['id'])
    print(f"  ✓ 共 {n_total} 条序列 | 重链(H): {n_heavy} | 轻链(L): {n_light} | 其他: {n_total - n_heavy - n_light}")
    for r in records[:5]:
        print(f"    {r['id']}  ({len(r['seq'])} nt)")

    # ════════════════════════════════════════════════════════════════
    #  Step 3 — IgBLAST
    # ════════════════════════════════════════════════════════════════
    print("\n" + "=" * 60)
    print("  Step 3 — 运行 IgBLAST")
    print("=" * 60)

    all_records = []
    for r in records:
        all_records.append({'id': f"{r['id']}||FWD", 'seq': r['seq']})
        all_records.append({'id': f"{r['id']}||RC",  'seq': reverse_complement(r['seq'])})

    print(f"  提交 {len(all_records)} 条序列（{n_total} 原始 × 2 方向）")
    print("  运行中，请稍候...")

    with tempfile.TemporaryDirectory() as tmpdir:
        fasta_path = os.path.join(tmpdir, 'input.fasta')
        write_fasta(all_records, fasta_path)
        airr_df = run_igblast(fasta_path)

    n_hit = (airr_df['v_call'].notna() & (airr_df['v_call'].str.strip() != '')).sum()
    print(f"  ✓ 完成！结果行: {len(airr_df)} | V 基因命中: {n_hit}")

    # ════════════════════════════════════════════════════════════════
    #  Step 4 — 提取可变区
    # ════════════════════════════════════════════════════════════════
    print("\n" + "=" * 60)
    print("  Step 4 — 提取可变区并质控")
    print("=" * 60)

    airr_map = {row['sequence_id']: row for _, row in airr_df.iterrows()}
    results = []

    for r in records:
        orig_id   = r['id']
        fwd_row   = airr_map.get(f"{orig_id}||FWD")
        rc_row    = airr_map.get(f"{orig_id}||RC")
        fwd_score = score_hit(fwd_row)
        rc_score  = score_hit(rc_row)

        if fwd_score == 0 and rc_score == 0:
            results.append({
                'sequence_id':         orig_id,
                'chain_type':          'N/A',
                'v_gene':              'N/A',
                'j_gene':              'N/A',
                'variable_region_dna': '', 'variable_region_aa': '',
                'vr_length_nt': 0, 'vr_length_aa': 0,
                'cdr1_dna': '', 'cdr1_aa': '', 'cdr1_length': 0,
                'cdr2_dna': '', 'cdr2_aa': '', 'cdr2_length': 0,
                'cdr3_dna': '', 'cdr3_aa': '', 'cdr3_length': 0,
                'orientation': 'N/A',
                'qc_status': 'FAIL_NO_HIT',
                'qc_detail': 'No V gene match in either orientation',
            })
            continue

        best_row, orientation = (fwd_row, 'forward') if fwd_score >= rc_score else (rc_row, 'reverse_complement')
        vr_dna = concat_vr(best_row, VR_DNA_COLS)
        vr_aa  = concat_vr(best_row, VR_AA_COLS)
        qc_status, qc_detail = qc_check(best_row, vr_dna, vr_aa)

        # 提取单独的 CDR 序列（按所选定义）
        if vr_aa and CDR_DEF not in CDR_IGBLAST_NATIVE:
            ab_result = extract_cdrs_abnumber(vr_aa, vr_dna, CDR_DEF)
            if ab_result:
                cdr1_dna, cdr1_aa, cdr1_len = ab_result[0], ab_result[1], ab_result[2]
                cdr2_dna, cdr2_aa, cdr2_len = ab_result[3], ab_result[4], ab_result[5]
                cdr3_dna, cdr3_aa, cdr3_len = ab_result[6], ab_result[7], ab_result[8]
            else:
                cdr1_dna = cdr2_dna = cdr3_dna = ''
                cdr1_aa  = cdr2_aa  = cdr3_aa  = ''
        else:
            cdr1_dna = clean_cdr(_get(best_row, 'cdr1'))
            cdr2_dna = clean_cdr(_get(best_row, 'cdr2'))
            cdr3_dna = clean_cdr(_get(best_row, 'cdr3'))
            cdr1_aa  = clean_cdr(_get(best_row, 'cdr1_aa'))
            cdr2_aa  = clean_cdr(_get(best_row, 'cdr2_aa'))
            cdr3_aa  = clean_cdr(_get(best_row, 'cdr3_aa'))

        results.append({
            'sequence_id':         orig_id,
            'chain_type':          _get(best_row, 'locus'),
            'v_gene':              _get(best_row, 'v_call'),
            'j_gene':              _get(best_row, 'j_call'),
            # 完整可变区
            'variable_region_dna': vr_dna,
            'variable_region_aa':  vr_aa,
            'vr_length_nt':        len(vr_dna),
            'vr_length_aa':        len(vr_aa),
            # CDR 单独列
            'cdr1_dna':           cdr1_dna,
            'cdr1_aa':            cdr1_aa,
            'cdr2_dna':           cdr2_dna,
            'cdr2_aa':            cdr2_aa,
            'cdr3_dna':           cdr3_dna,
            'cdr3_aa':            cdr3_aa,
            'cdr1_length':         len(cdr1_dna),
            'cdr2_length':         len(cdr2_dna),
            'cdr3_length':         len(cdr3_dna),
            # 方向 & 质控
            'orientation':         orientation,
            'qc_status':           qc_status,
            'qc_detail':           qc_detail,
        })

    df_all = pd.DataFrame(results)
    df_pass = df_all[df_all['qc_status'].str.startswith('PASS')].copy()
    df_fail = df_all[~df_all['qc_status'].str.startswith('PASS')].copy()

    print(f"  ✓ 提取完成 | 总: {len(df_all)} | PASS: {len(df_pass)} ({len(df_pass)/len(df_all)*100:.1f}%) | FAIL: {len(df_fail)}")
    print("\n  QC 状态分布:")
    print(df_all['qc_status'].value_counts().to_string())

    # ════════════════════════════════════════════════════════════════
    #  Step 5 — 输出 Excel
    # ════════════════════════════════════════════════════════════════
    print("\n" + "=" * 60)
    print("  Step 5 — 保存 Excel")
    print("=" * 60)

    # 定义并排序输出列
    output_columns = [
        'sequence_id', 'chain_type', 'v_gene', 'j_gene',
        'variable_region_dna', 'variable_region_aa',
        'vr_length_nt', 'vr_length_aa',
        'cdr1_dna', 'cdr1_aa', 'cdr1_length',
        'cdr2_dna', 'cdr2_aa', 'cdr2_length',
        'cdr3_dna', 'cdr3_aa', 'cdr3_length',
        'orientation', 'qc_status', 'qc_detail',
    ]
    output_columns = [c for c in output_columns if c in df_pass.columns]
    df_pass = df_pass[output_columns]
    df_fail = df_fail[output_columns]

    dna_cols = [
        'variable_region_dna', 'variable_region_aa',
        'cdr1_dna', 'cdr1_aa', 'cdr2_dna', 'cdr2_aa', 'cdr3_dna', 'cdr3_aa',
    ]

    with pd.ExcelWriter(OUTPUT_XLSX, engine='openpyxl') as writer:
        pass_sheet_name = f'PASS_{CDR_DEF.upper()}'
        fail_sheet_name = f'FAIL_{CDR_DEF.upper()}'
        df_pass.to_excel(writer, sheet_name=pass_sheet_name, index=False, na_rep='')
        df_fail.to_excel(writer, sheet_name=fail_sheet_name, index=False, na_rep='')
        wb = writer.book
        style_sheet(wb[pass_sheet_name], df_pass, hdr_color="1F6B3A", alt_color="EAF4EC",
                    dna_cols=dna_cols)
        style_sheet(wb[fail_sheet_name], df_fail, hdr_color="8B1A1A", alt_color="FDF0F0")

    print(f"  ✓ Excel 已保存: {OUTPUT_XLSX}")
    print(f"     PASS sheet: {len(df_pass)} 条")
    print(f"     FAIL sheet: {len(df_fail)} 条")

    return records, df_all


# ═══════════════════════════════════════════════════════════════════
#  Excel 样式函数（模块级，避免重复定义）
# ═══════════════════════════════════════════════════════════════════
def style_sheet(ws, df, hdr_color, alt_color, dna_cols=None, row_height=60):
    """通用 Excel 工作表样式"""
    col_widths = {
        'sequence_id': 40, 'chain_type': 12, 'v_gene': 30, 'j_gene': 10,
        'variable_region_dna': 65, 'variable_region_aa': 65,
        'vr_length_nt': 14, 'vr_length_aa': 14,
        'cdr1_dna': 50, 'cdr1_aa': 50, 'cdr1_length': 14,
        'cdr2_dna': 50, 'cdr2_aa': 50, 'cdr2_length': 14,
        'cdr3_dna': 50, 'cdr3_aa': 50, 'cdr3_length': 14,
        'orientation': 22, 'qc_status': 24, 'qc_detail': 40,
    }
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 18)
    ws.row_dimensions[1].height = 24
    for i in range(2, len(df) + 2):
        ws.row_dimensions[i].height = row_height

    hdr_fill = PatternFill("solid", fgColor=hdr_color)
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(bottom=Side(style='medium', color='AAAAAA'))

    dna_col_idxs = set()
    if dna_cols:
        for dc in dna_cols:
            col_list = list(df.columns)
            if dc in col_list:
                dna_col_idxs.add(col_list.index(dc) + 1)

    thin = Border(
        left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),  bottom=Side(style='thin', color='DDDDDD'),
    )
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=len(df) + 1), 1):
        fill = PatternFill("solid", fgColor=alt_color) if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for cell in row:
            cell.fill = fill
            cell.border = thin
            if cell.column in dna_col_idxs:
                cell.font = Font(name='Courier New', size=9)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            else:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions


STATUS_COLORS = {'PAIRED_PASS': 'D4EDDA', 'PAIRED_PARTIAL': 'FFF3CD',
                 'VH_ONLY': 'D1ECF1', 'VL_ONLY': 'D1ECF1'}


def style_paired_sheet(ws, df, hdr_color, row_height=28):
    """配对结果工作表样式"""
    for col_idx in range(1, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16
    ws.row_dimensions[1].height = 24
    for i in range(2, len(df) + 2):
        ws.row_dimensions[i].height = row_height

    hdr_fill = PatternFill("solid", fgColor=hdr_color)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(bottom=Side(style='medium', color='888888'))

    pair_col_idx = list(df.columns).index('pair_status') + 1 if 'pair_status' in df.columns else None
    thin = Border(left=Side(style='thin',color='DDDDDD'), right=Side(style='thin',color='DDDDDD'),
                  top=Side(style='thin',color='DDDDDD'),  bottom=Side(style='thin',color='DDDDDD'))

    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=len(df)+1), 1):
        pair_val = str(row[pair_col_idx-1].value or '') if pair_col_idx else ''
        row_color = STATUS_COLORS.get(pair_val, 'F5F5F5' if i % 2 == 0 else 'FFFFFF')
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=row_color)
            cell.border = thin
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions


# ═══════════════════════════════════════════════════════════════════
#  SECTION 3 — COS 上传 & 下载
# ═══════════════════════════════════════════════════════════════════
def cos_upload_input():
    """上传输入序列文件到 COS"""
    if not COS_ENABLED:
        return

    print("\n" + "-" * 50)
    print("  [COS] 上传输入文件...")
    cos_key = f'{COS_PREFIX}input/{Path(INPUT_FILE).name}'
    try:
        with open(INPUT_FILE, 'rb') as f:
            COS_CLIENT.put_object(Bucket=COS_BUCKET, Body=f, Key=cos_key, StorageClass='STANDARD')
        print(f"  ✅ 已上传输入文件: cos://{COS_BUCKET}/{cos_key}")
    except Exception as e:
        print(f"  ⚠️  上传失败: {e}")


def cos_upload_results():
    """上传分析结果到 COS"""
    if not COS_ENABLED:
        return

    print("\n" + "-" * 50)
    print("  [COS] 上传结果文件...")
    files_to_upload = [OUTPUT_XLSX]
    paired = OUTPUT_XLSX.replace('.xlsx', '_VH_VL_paired.xlsx')
    if os.path.exists(paired):
        files_to_upload.append(paired)

    for local_path in files_to_upload:
        cos_key = f'{COS_PREFIX}output/{Path(local_path).name}'
        try:
            with open(local_path, 'rb') as f:
                COS_CLIENT.put_object(Bucket=COS_BUCKET, Body=f, Key=cos_key, StorageClass='STANDARD')
            print(f"  ✅ 已上传: cos://{COS_BUCKET}/{cos_key}")
        except Exception as e:
            print(f"  ⚠️  上传失败 {local_path}: {e}")

    print("  📤 COS 上传完成")


def cos_download_results():
    """从 COS 下载结果文件到本地"""
    if not COS_ENABLED:
        return

    print("\n" + "-" * 50)
    print("  [COS] 可用结果文件列表:")
    prefix = f'{COS_PREFIX}output/'
    try:
        resp = COS_CLIENT.list_objects(Bucket=COS_BUCKET, Prefix=prefix, Delimiter='/')
        contents = resp.get('Contents', [])
        if not contents:
            print("  (暂无结果文件)")
            return

        files = [c['Key'] for c in contents if c['Key'] != prefix]
        if not files:
            print("  (暂无结果文件)")
            return

        for i, f in enumerate(files):
            print(f"  [{i+1}] cos://{COS_BUCKET}/{f}")

        choice = input("\n  输入编号下载到本地 (回车跳过): ").strip()
        if not choice.isdigit():
            return
        idx = int(choice) - 1
        if idx < 0 or idx >= len(files):
            return

        remote_key = files[idx]
        local_name = Path(remote_key).name
        local_path = os.path.expanduser(f'~/antibody/{local_name}')

        COS_CLIENT.download_file(Bucket=COS_BUCKET, Key=remote_key, DestFilePath=local_path)
        size = os.path.getsize(local_path)
        print(f"  ✅ 已下载: {local_path} ({size:,} bytes)")

    except Exception as e:
        print(f"  ⚠️  COS 操作失败: {e}")


def cos_menu():
    """COS 功能菜单：上传/下载"""
    if not COS_ENABLED:
        print("\n  ⏭  COS 未启用，跳过")
        return

    print("\n" + "=" * 60)
    print("  Step 6 — COS 上传 / 下载")
    print("=" * 60)
    print("  1) 上传输入文件到 COS")
    print("  2) 上传结果到 COS")
    print("  3) 从 COS 下载结果")
    print("  4) 全部跳过")

    choice = input("\n  请选择 [默认: 2]: ").strip() or '2'

    if choice == '1':
        cos_upload_input()
    elif choice == '2':
        cos_upload_results()
    elif choice == '3':
        cos_download_results()
    elif choice == '4':
        print("  ⏭  跳过 COS 操作")
    else:
        print("  ⏭  跳过 COS 操作")


# ═══════════════════════════════════════════════════════════════════
#  SECTION 4 — VH/VL 配对（可选）
# ═══════════════════════════════════════════════════════════════════
def pair_vh_vl(records, df_all):
    """VH/VL 配对分析"""
    print("\n" + "=" * 60)
    print("  Step 7 — VH/VL 配对（可选）")
    print("=" * 60)

    pair_yes = input("\n  是否进行 VH/VL 配对? (y/N): ").strip().lower()
    if pair_yes != 'y':
        print("  ⏭  跳过配对")
        return False

    # 提取孔位编号
    def extract_well(seq_id):
        m = re.search(r'-B\d+-(\d+[A-Z]\d+)-[HL]-', seq_id)
        return m.group(1) if m else None

    df_all['well'] = df_all['sequence_id'].apply(extract_well)
    df_heavy = df_all[df_all['sequence_id'].str.contains('-H-')].copy()
    df_light = df_all[df_all['sequence_id'].str.contains('-L-')].copy()

    all_wells = sorted(set(df_heavy['well'].dropna()) | set(df_light['well'].dropna()))
    print(f"  ✓ 重链孔位: {df_heavy['well'].nunique()} 个")
    print(f"    轻链孔位: {df_light['well'].nunique()} 个")
    print(f"    总孔位:   {len(all_wells)} 个")

    CHAIN_KEYS = (['sequence_id', 'chain_type', 'v_gene', 'j_gene',
                   'variable_region_dna', 'variable_region_aa',
                   'vr_length_nt', 'vr_length_aa', 'qc_status', 'qc_detail']
                  + CDR_COLS)

    def get_chain_info(df_chain, well, prefix):
        rows = df_chain[df_chain['well'] == well]
        if len(rows) == 0:
            return {f'{prefix}_{k}': '' for k in CHAIN_KEYS}
        return {f'{prefix}_{k}': rows.iloc[0][k] for k in CHAIN_KEYS}

    paired_rows = []
    for well in all_wells:
        h_info = get_chain_info(df_heavy, well, 'VH')
        l_info = get_chain_info(df_light, well, 'VL')
        has_h  = bool(h_info['VH_sequence_id'])
        has_l  = bool(l_info['VL_sequence_id'])
        h_pass = has_h and str(h_info['VH_qc_status']).startswith('PASS')
        l_pass = has_l and str(l_info['VL_qc_status']).startswith('PASS')

        if   h_pass and l_pass:  pair_status = 'PAIRED_PASS'
        elif has_h  and has_l:   pair_status = 'PAIRED_PARTIAL'
        elif has_h:              pair_status = 'VH_ONLY'
        elif has_l:              pair_status = 'VL_ONLY'
        else:                    pair_status = 'MISSING'

        row = {'well': well, 'pair_status': pair_status}
        row.update(h_info)
        row.update(l_info)
        paired_rows.append(row)

    df_paired = pd.DataFrame(paired_rows)

    print("\n  配对状态统计:")
    print(df_paired['pair_status'].value_counts().to_string())

    # ── 输出配对 Excel ────────────────────────────────────────────
    PAIRED_OUTPUT = OUTPUT_XLSX.replace('.xlsx', '_VH_VL_paired.xlsx')

    overview_cols = [
        'well', 'pair_status',
        'VH_sequence_id', 'VH_v_gene', 'VH_j_gene', 'VH_vr_length_nt', 'VH_vr_length_aa', 'VH_qc_status',
        'VL_sequence_id', 'VL_v_gene', 'VL_j_gene', 'VL_vr_length_nt', 'VL_vr_length_aa', 'VL_qc_status',
    ]
    full_cols = [
        'well', 'pair_status',
        'VH_sequence_id', 'VH_v_gene', 'VH_j_gene', 'VH_variable_region_dna', 'VH_variable_region_aa',
        'VH_vr_length_nt', 'VH_vr_length_aa', 'VH_qc_status',
    ]
    # 添加 VH CDR 列
    for c in CDR_COLS:
        full_cols.append(f'VH_{c}')
    full_cols += [
        'VL_sequence_id', 'VL_v_gene', 'VL_j_gene', 'VL_variable_region_dna', 'VL_variable_region_aa',
        'VL_vr_length_nt', 'VL_vr_length_aa', 'VL_qc_status',
    ]
    for c in CDR_COLS:
        full_cols.append(f'VL_{c}')

    df_overview = df_paired[overview_cols]
    df_full     = df_paired[df_paired['pair_status'] == 'PAIRED_PASS'][full_cols]
    df_unpaired = df_paired[df_paired['pair_status'] != 'PAIRED_PASS'][overview_cols]

    with pd.ExcelWriter(PAIRED_OUTPUT, engine='openpyxl') as writer:
        df_overview.to_excel(writer, sheet_name='配对总览',        index=False, na_rep='')
        df_full.to_excel(writer, sheet_name='PAIRED_PASS序列',    index=False, na_rep='')
        df_unpaired.to_excel(writer, sheet_name='未配对或FAIL',    index=False, na_rep='')
        wb = writer.book
        style_paired_sheet(wb['配对总览'],     df_overview, "2C5F8A")
        style_paired_sheet(wb['PAIRED_PASS序列'], df_full, "1F6B3A", row_height=40)
        style_paired_sheet(wb['未配对或FAIL'], df_unpaired, "8B4513")

    print(f"\n  ✓ 配对 Excel 已保存: {PAIRED_OUTPUT}")
    print(f"     配对总览: {len(df_overview)} 个孔位")
    print(f"     PAIRED_PASS: {len(df_full)} 对")
    print(f"     未配对: {len(df_unpaired)} 个")

    # 自动上传配对结果到 COS
    if COS_ENABLED:
        cos_key = f'{COS_PREFIX}output/{Path(PAIRED_OUTPUT).name}'
        try:
            with open(PAIRED_OUTPUT, 'rb') as f:
                COS_CLIENT.put_object(Bucket=COS_BUCKET, Body=f, Key=cos_key, StorageClass='STANDARD')
            print(f"  📤 配对结果已上传: cos://{COS_BUCKET}/{cos_key}")
        except Exception as e:
            print(f"  ⚠️  上传失败: {e}")


# ═══════════════════════════════════════════════════════════════════
#  SECTION 5 — 进化树分析（合并 phylo_tree 的改进）
# ═══════════════════════════════════════════════════════════════════

# V 基因家族 → 颜色
VH_FAMILY_COLORS = [
    '#1f77b4','#ff7f0e','#2ca02c','#d62728','#9467bd',
    '#8c564b','#e377c2','#7f7f7f','#bcbd22','#17becf',
    '#aec7e8','#ffbb78','#98df8a','#ff9896','#c5b0d5',
]
J_MARKERS = {
    'IGHJ1':'o','IGHJ2':'s','IGHJ3':'^','IGHJ4':'D',
    'IGHJ5':'P','IGHJ6':'X','IGHJ': 'o',
}
J_MARKER_DEFAULT = 'P'


def _get_vgene_family(vgene_str):
    """从 V 基因调用中提取家族名（如 IGHV1-2*01 → IGHV1）"""
    if not vgene_str or str(vgene_str).lower() in ('nan','','n/a'):
        return 'Unknown'
    first = str(vgene_str).split(',')[0].strip().split('*')[0]
    m = re.match(r'(IG[HKL]V\d+)', first)
    return m.group(1) if m else first.split('-')[0] if '-' in first else first


def _compute_p_distance(aligned_seqs, names):
    """计算 p-distance 矩阵（gap-free 位点上的差异率）"""
    N = len(aligned_seqs)
    dm = np.zeros((N, N))
    identical_pairs = []
    for i in range(N):
        for j in range(i+1, N):
            diff = shared = 0
            for a, b in zip(aligned_seqs[i], aligned_seqs[j]):
                if a == '-' or b == '-': continue
                shared += 1
                if a != b: diff += 1
            d = diff / shared if shared > 0 else 1.0
            if d == 0:
                d = 1e-6
                identical_pairs.append((names[i], names[j]))
            dm[i][j] = dm[j][i] = d
    return dm, identical_pairs


def _build_midpoint_tree(tree):
    """中点生根"""
    if tree.rooted: return
    tips = tree.get_terminals()
    if not tips: return
    # 找最远叶节点对
    max_dist = -1
    far_pair = (None, None)
    for i, t1 in enumerate(tips):
        for t2 in tips[i+1:]:
            d = tree.distance(t1, t2)
            if d > max_dist:
                max_dist = d
                far_pair = (t1, t2)
    if far_pair[0] is None: return
    # 找中点
    dist_to_far = tree.distance(far_pair[0], tree.root)
    half = max_dist / 2
    target = dist_to_far - half
    # 回溯找中点
    node = far_pair[0]
    path = [node]
    while node != tree.root:
        node = tree.root if node.is_parent_of(tree.root) else next(tree.root.find_clades())
        break  # 简化：直接用 Phylo 的内置方法
    tree.root_at_midpoint()


def _draw_beautiful_tree(tree, paired_pass, identical_pairs, label, tree_dir):
    """
    全标注进化树可视化
    颜色 = VH V 基因家族，形状 = VH J 基因
    label = 'VH' 或 'VL'，决定用哪个序列 ID 作为 meta 查找键
    """
    # 用序列 ID (截断到50字符，与 MUSCLE FASTA ID 一致) 建立 meta 映射
    id_key = 'VH_sequence_id' if label == 'VH' else 'VL_sequence_id'
    meta = {}
    for r in paired_pass:
        sid = str(r.get(id_key, '') or '')[:50]
        if sid:
            meta[sid] = r

    # 提取 V 基因家族（优先用 VH，如果 label=VL 则用 VL）
    fam_key = 'VH_v_family' if label == 'VH' else 'VL_v_family'
    vgene_key = 'VH_v_gene' if label == 'VH' else 'VL_v_gene'
    jgene_key = 'VH_j_gene' if label == 'VH' else 'VL_j_gene'

    families = sorted(set(
        _get_vgene_family(str(r.get(vgene_key, '')))
        for r in paired_pass
    ) or ['Unknown'])
    fam_color = {fam: VH_FAMILY_COLORS[i % len(VH_FAMILY_COLORS)]
                 for i, fam in enumerate(families)}

    identical_wells = set()
    for a, b in identical_pairs:
        identical_wells.add(a); identical_wells.add(b)

    N = tree.count_terminals()
    fig_h = max(10, N * 0.25)
    fig_w = 18
    import matplotlib; matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    from matplotlib.lines import Line2D

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.set_facecolor('#FAFAFA')

    def _get_xy(tree):
        xp, yp = {}, {}
        leaves = tree.get_terminals()
        yp.update({leaf: i for i, leaf in enumerate(leaves)})
        def _walk(c, cumlen):
            xp[c] = cumlen
            for child in c.clades:
                bl = child.branch_length or 0
                _walk(child, cumlen + bl)
        _walk(tree.root, 0)
        x_max = max(xp.values()) or 1
        xp = {k: v/x_max for k, v in xp.items()}
        def _walk_y(c):
            if c.is_terminal(): return yp[c]
            ys = [_walk_y(ch) for ch in c.clades]
            yp[c] = sum(ys)/len(ys)
            return yp[c]
        _walk_y(tree.root)
        return xp, yp, x_max
    xp, yp, x_max = _get_xy(tree)

    def _draw_branch(c):
        x1, y1 = xp[c], yp[c]
        for ch in c.clades:
            x2, y2 = xp[ch], yp[ch]
            ax.plot([x1,x2],[y2,y2],color='#444',lw=0.7,zorder=1)
            ax.plot([x1,x1],[y1,y2],color='#444',lw=0.7,zorder=1)
            _draw_branch(ch)
    _draw_branch(tree.root)

    for leaf in tree.get_terminals():
        leaf_key = leaf.name
        r = meta.get(leaf_key, {})

        # 获取当前链的基因信息（过滤 nan/None/空值）
        def _gene(g):
            v = str(g).split(',')[0].split('*')[0] if g else ''
            return v if v and v.lower() not in ('nan','none','') else '?'
        vh_v = _gene(r.get('VH_v_gene'))
        vh_j = _gene(r.get('VH_j_gene'))
        vl_v = _gene(r.get('VL_v_gene'))
        vl_j = _gene(r.get('VL_j_gene'))

        # 当前链的 V 基因家族
        cur_vgene = str(r.get(vgene_key, ''))
        vfam = _get_vgene_family(cur_vgene) if cur_vgene else 'Unknown'

        # 当前链的 J 基因（用于形状编码）
        cur_jgene_raw = str(r.get(jgene_key, ''))
        cur_jgene = cur_jgene_raw.split(',')[0].split('*')[0] if cur_jgene_raw and cur_jgene_raw.lower() not in ('nan','none','') else ''

        color = fam_color.get(vfam, '#999')
        marker = J_MARKERS.get(cur_jgene, J_MARKER_DEFAULT)

        xv, yv = xp[leaf], yp[leaf]
        ax.scatter(xv, yv, c=color, marker=marker, s=50, zorder=3,
                   edgecolors='white', linewidths=0.5)

        # 标签：孔位 | 双链基因信息
        well_label = str(r.get('well', leaf_key[:12]))
        label_text = f"{well_label}  VH:{vh_v} {vh_j}  VL:{vl_v} {vl_j}"
        ax.text(xv+0.008, yv, label_text, va='center', ha='left',
                fontsize=6.5, color='#222', fontfamily='monospace')

        if leaf_key in identical_wells:
            ax.text(xv+0.008+0.33, yv, '* identical clone',
                    va='center', ha='left', fontsize=6, color='#C00', style='italic')

    # V 基因图例（使用当前链的 V 家族）
    legend_color = []
    for fam in sorted(fam_color.keys()):
        cnt = 0
        for r in paired_pass:
            v = str(r.get(vgene_key, ''))
            if v.lower() not in ('nan', 'none', ''):
                if _get_vgene_family(v) == fam:
                    cnt += 1
        legend_color.append(mpatches.Patch(color=fam_color[fam], label=f"{fam} (n={cnt})"))
    leg1 = ax.legend(handles=legend_color,
                     title=f'{label} V-gene family',
                     title_fontsize=8, fontsize=7, loc='lower left',
                     bbox_to_anchor=(0.0,0.0), framealpha=0.92, edgecolor='#CCC')

    # J 基因图例
    def _j_name(r):
        v = str(r.get(jgene_key, ''))
        if v.lower() in ('nan', 'none', '', '?'):
            return None
        return v.split(',')[0].split('*')[0]
    j_present = sorted(set(
        _j_name(r) for r in paired_pass if _j_name(r)))
    j_items = []
    for j in j_present:
        cnt = sum(1 for r in paired_pass if _j_name(r) == j)
        mk = J_MARKERS.get(j, J_MARKER_DEFAULT)
        j_items.append(Line2D([0],[0], marker=mk, color='w',
                              markerfacecolor='#555', markersize=8, label=f"{j} (n={cnt})"))
    leg2 = ax.legend(handles=j_items, title=f'{label} J-gene',
                     title_fontsize=8, fontsize=7, loc='lower left',
                     bbox_to_anchor=(0.0,0.13), framealpha=0.92, edgecolor='#CCC')
    ax.add_artist(leg1)

    ax.set_xlim(-0.02, 1.55)
    ax.set_ylim(-1, N)
    ax.set_xlabel('p-distance (normalized)', fontsize=9)
    ax.set_title(f'{label} Phylogenetic Tree | n={N} | NJ / midpoint-rooted',
                 fontsize=11, pad=12)
    for spine in ['top','right','left']: ax.spines[spine].set_visible(False)
    ax.set_yticks([])
    xticks = np.linspace(0, 1, 6)
    ax.set_xticks(xticks)
    ax.set_xticklabels([f'{v*x_max:.3f}' for v in xticks], fontsize=7)

    plt.tight_layout(pad=1.5)
    svg_path = os.path.join(tree_dir, f'{label}_tree.svg')
    png_path = os.path.join(tree_dir, f'{label}_tree.png')
    fig.savefig(svg_path, format='svg', bbox_inches='tight')
    fig.savefig(png_path, format='png', bbox_inches='tight', dpi=200)
    plt.close()
    print(f"    SVG:   {svg_path}")
    print(f"    PNG:   {png_path}")
    return svg_path, png_path


def build_phylogenetic_trees():
    """对配对 VH/VL 构建 NJ 树（MUSCLE + p-distance + 中点生根 + SVG/PNG）"""
    paired_excel = OUTPUT_XLSX.replace('.xlsx', '_VH_VL_paired.xlsx')
    if not os.path.exists(paired_excel):
        print("\n  ⏭  未找到配对文件，跳过进化树分析")
        return False
    print("\n" + "=" * 60)
    print("  Step 8 — 进化树分析（NJ / p-distance / midpoint-root）")
    print("=" * 60)

    df_paired = pd.read_excel(paired_excel, sheet_name='PAIRED_PASS序列')
    if len(df_paired) < 3:
        print(f"  ⏭  PAIRED_PASS 仅 {len(df_paired)} 条")
        return False
    print(f"  ✓ 读取 {len(df_paired)} 对 PAIRED_PASS 序列")

    tree_dir = f'{WORK_DIR}/phylogenetic_trees'
    os.makedirs(tree_dir, exist_ok=True)

    import subprocess, tempfile
    from Bio import SeqIO, AlignIO, Phylo
    from Bio.Seq import Seq
    from Bio.SeqRecord import SeqRecord
    from Bio.Phylo.TreeConstruction import DistanceMatrix, DistanceTreeConstructor

    def _build_chain(seqs, ids, label):
        """为单链构建 NJ 树 + 可视化"""
        valid = [(s, sid) for s, sid in zip(seqs, ids)
                 if isinstance(s, str) and len(s) >= 20]
        if len(valid) < 3:
            print(f"\n  [{label}] 有效序列 {len(valid)} 条，跳过")
            return
        clean_seqs, clean_ids = zip(*valid)
        print(f"\n  [{label}] {len(clean_seqs)} 条序列")

        with tempfile.TemporaryDirectory() as tmp:
            fa_in = os.path.join(tmp, f'{label}.fasta')
            fa_aln = os.path.join(tmp, f'{label}_aln.fasta')
            SeqIO.write([SeqRecord(Seq(s), id=sid[:50], description='')
                         for s, sid in zip(clean_seqs, clean_ids)], fa_in, 'fasta')
            print(f"    MUSCLE 对齐...")
            r = subprocess.run(['muscle','-align',fa_in,'-output',fa_aln],
                               capture_output=True, text=True)
            if r.returncode != 0:
                print(f"    ❌ MUSCLE 失败: {r.stderr[:150]}")
                return
            aln = AlignIO.read(fa_aln, 'fasta')
            aln_len = aln.get_alignment_length()
            print(f"    对齐长度: {aln_len} aa")

            # p-distance
            seq_strs = [str(s.seq) for s in aln]
            names = [s.id for s in aln]
            dm, identical_pairs = _compute_p_distance(seq_strs, names)

            # NJ 树
            N = len(names)
            dm_lower = [[dm[i][j] for j in range(i+1)] for i in range(N)]
            tree = DistanceTreeConstructor().nj(DistanceMatrix(names, dm_lower))
            try: tree.root_at_midpoint()
            except: pass
            for cl in tree.find_clades():
                if cl.name: cl.name = cl.name[:40].replace(' ','_')

            nwk = os.path.join(tree_dir, f'{label}_tree.nwk')
            Phylo.write(tree, nwk, 'newick')
            print(f"    Newick: {nwk}")

            # 原始配对数据（用于美化绘图）
            paired_pass = df_paired.to_dict('records')
            for r in paired_pass:
                if r.get('well') is None:
                    r['well'] = r.get('VH_sequence_id','')[:10]
            _draw_beautiful_tree(tree, paired_pass, identical_pairs, label, tree_dir)

    vh_seqs = df_paired.get('VH_variable_region_aa', pd.Series(dtype=str)).tolist()
    vh_ids  = df_paired.get('VH_sequence_id', pd.Series(dtype=str)).tolist()
    vl_seqs = df_paired.get('VL_variable_region_aa', pd.Series(dtype=str)).tolist()
    vl_ids  = df_paired.get('VL_sequence_id', pd.Series(dtype=str)).tolist()

    _build_chain(vh_seqs, vh_ids, 'VH')
    _build_chain(vl_seqs, vl_ids, 'VL')

    print(f"\n  ✅ 进化树分析完成! 结果目录: {tree_dir}")
    return True


# ═══════════════════════════════════════════════════════════════════
#  MAIN — 支持命令行参数（合并自 phylo_tree.py）
# ═══════════════════════════════════════════════════════════════════

def run_interactive():
    """交互式运行（默认模式）"""
    if not step0_install(): return
    if not set_config(): return
    records, df_all = analyze()
    cos_menu()
    pair_vh_vl(records, df_all)
    build_phylogenetic_trees()
    _print_summary()


def run_cli(args):
    """命令行模式（非交互式）"""
    global INPUT_FILE, ORGANISM, CDR_DEF, OUTPUT_XLSX, AUX_FILE
    INPUT_FILE = args.input
    if args.species: ORGANISM = args.species
    if args.cdr_def: CDR_DEF = args.cdr_def
    OUTPUT_XLSX = f'{WORK_DIR}/antibody_results_{CDR_DEF}.xlsx'
    AUX_FILE = f'{IGBLAST_DIR}/optional_file/{ORGANISM}_gl.aux'
    print(f"[CLI] {INPUT_FILE} | {ORGANISM} | CDR={CDR_DEF}")

    if not args.skip_setup and not step0_install():
        print("❌ 环境安装失败"); return
    if not os.path.exists(INPUT_FILE):
        print(f"❌ 文件不存在: {INPUT_FILE}"); return

    records, df_all = analyze()
    if args.pair:
        # 命令行模式下自动跳过交互提示
        import builtins
        _orig_input = builtins.input
        builtins.input = lambda *a, **kw: 'y'
        try:
            pair_vh_vl(records, df_all)
        finally:
            builtins.input = _orig_input
    if args.tree: build_phylogenetic_trees()
    _print_summary()


def _print_summary():
    print("\n" + "=" * 60)
    print("  🎉 全部完成！")
    print(f"  结果文件: {OUTPUT_XLSX}")
    paired = OUTPUT_XLSX.replace('.xlsx', '_VH_VL_paired.xlsx')
    if os.path.exists(paired): print(f"  配对文件: {paired}")
    tree_dir = f'{WORK_DIR}/phylogenetic_trees'
    if os.path.isdir(tree_dir) and any(f.endswith('.nwk') for f in os.listdir(tree_dir)):
        print(f"  进化树: {tree_dir}/")
    print("=" * 60)


def main():
    parser = argparse.ArgumentParser(
        description='抗体可变区序列提取 + 进化树分析',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__)
    parser.add_argument('positional', nargs='?', help='输入 FASTA/FASTQ 文件路径')
    parser.add_argument('--input', help='输入文件路径（命令行模式）')
    parser.add_argument('--species', default='mouse',
                        choices=['mouse','human','rabbit','rat'], help='物种')
    parser.add_argument('--cdr_def', default='kabat',
                        choices=['kabat','imgt','chothia','contact','martin'],
                        help='CDR 定义')
    parser.add_argument('--skip_setup', action='store_true', help='跳过环境安装')
    parser.add_argument('--pair', action='store_true', help='进行 VH/VL 配对')
    parser.add_argument('--tree', action='store_true', help='构建进化树')
    args = parser.parse_args()

    # 命令行模式（--input 或位置参数 + --pair/--tree）
    if args.input or (args.positional and (args.pair or args.tree)):
        if not args.input: args.input = args.positional
        run_cli(args)
    elif args.positional:
        # 交互式模式 + 位置参数作为输入文件
        sys.argv = [sys.argv[0], args.positional]
        run_interactive()
    else:
        # 纯交互式
        run_interactive()


if __name__ == '__main__':
    main()
